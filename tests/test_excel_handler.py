"""Testes para o módulo excel_handler."""

from pathlib import Path

import openpyxl
import pytest

from inventory_count_automation.excel_handler import assign_balances


@pytest.fixture
def sample_workbook(tmp_path: Path) -> tuple[openpyxl.Workbook, Path]:
    """Cria um workbook de teste simulando a planilha pré-preenchida."""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook sem planilha ativa.")
    ws.title = "Inventário_Empresa 1"

    # Cabeçalhos na linha 2
    headers = {
        "B": "Empresa",
        "C": "SKU",
        "D": "Descrição",
        "E": "Posição",
        "F": "Depósito",
        "G": "Barcode",
        "H": "Volume/Série",
        "M": "QTD Físico",
    }
    for col, value in headers.items():
        ws[f"{col}2"] = value

    # Dados pré-preenchidos (linhas 3-7)
    products = [
        {"B": 1, "C": "SKU001", "D": "Produto A", "G": "MCS000PROD001"},
        {"B": 1, "C": "SKU002", "D": "Produto B", "G": "MCS000PROD002"},
        {"B": 1, "C": "SKU003", "D": "Produto C", "G": "MCS000PROD003"},
        {"B": 1, "C": "SKU004", "D": "Produto D", "G": "MCS000PROD004"},
        {"B": 1, "C": "SKU005", "D": "Produto E", "G": "MCS000PROD005"},
    ]
    for i, prod in enumerate(products, start=3):
        for col, value in prod.items():
            ws[f"{col}{i}"] = value

    # Salva num arquivo temporário (simula a planilha original)
    planilha_path = tmp_path / "planilha.xlsx"
    wb.save(planilha_path)

    # Recarrega para simular o fluxo real
    wb = openpyxl.load_workbook(planilha_path)
    return wb, planilha_path


class TestAssignBalances:
    def test_assigns_correct_quantities(self, sample_workbook) -> None:
        wb, planilha_path = sample_workbook
        counted = {
            "MCS000PROD001": 5,
            "MCS000PROD003": 12,
        }

        result = assign_balances(counted, wb=wb, save_path=planilha_path)

        assert "MCS000PROD001" in result["matched"]
        assert "MCS000PROD003" in result["matched"]
        assert len(result["not_found"]) == 0

        # Verifica os valores escritos na planilha (recarrega do disco)
        wb_check = openpyxl.load_workbook(planilha_path)
        ws = wb_check.active
        if ws is None:
            raise RuntimeError("Workbook sem planilha ativa.")
        assert ws["M3"].value == 5   # PROD001
        assert ws["M5"].value == 12  # PROD003
        # Produtos não contados devem permanecer sem valor
        assert ws["M4"].value is None  # PROD002

    def test_reports_not_found_barcodes(self, sample_workbook) -> None:
        wb, planilha_path = sample_workbook
        counted = {
            "MCS000PROD001": 3,
            "MCS000FANTASMA": 7,
        }

        result = assign_balances(counted, wb=wb, save_path=planilha_path)

        assert "MCS000PROD001" in result["matched"]
        assert "MCS000FANTASMA" in result["not_found"]

    def test_empty_counted(self, sample_workbook) -> None:
        wb, planilha_path = sample_workbook
        result = assign_balances({}, wb=wb, save_path=planilha_path)

        assert result["matched"] == []
        assert result["not_found"] == []

    def test_case_insensitive_matching(self, sample_workbook) -> None:
        wb, planilha_path = sample_workbook
        counted = {"MCS000PROD002": 10}

        result = assign_balances(counted, wb=wb, save_path=planilha_path)
        assert "MCS000PROD002" in result["matched"]
