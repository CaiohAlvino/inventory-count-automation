"""Identificação dos produtos na planilha e atribuição dos saldos contados."""

from pathlib import Path

import openpyxl

from inventory_count_automation.config import (
    COL_BARCODE,
    COL_QTD_FISICO,
    DATA_START_ROW,
    INPUT_PLANILHA_DIR,
    PLANILHA_BASE_FILENAME,
)


def _build_barcode_index(ws, col_barcode: str, start_row: int) -> dict[str, int]:
    """
    Percorre a coluna de barcode da planilha e cria um índice
    {barcode_upper: número_da_linha} para busca O(1).
    """
    index: dict[str, int] = {}

    for row in range(start_row, ws.max_row + 1):
        cell_value = ws[f"{col_barcode}{row}"].value
        if cell_value is not None:
            barcode = str(cell_value).strip().upper()
            if barcode:
                index[barcode] = row

    return index


def _default_planilha_path() -> Path:
    """Retorna o caminho padrão da planilha base."""
    return INPUT_PLANILHA_DIR / PLANILHA_BASE_FILENAME


def load_workbook(filepath: Path | None = None) -> tuple[openpyxl.Workbook, Path]:
    """Carrega a planilha base e retorna (workbook, caminho_do_arquivo)."""
    if filepath is None:
        filepath = _default_planilha_path()

    if not filepath.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {filepath}")

    return openpyxl.load_workbook(filepath), filepath


def assign_balances(
    counted: dict[str, int],
    wb: openpyxl.Workbook | None = None,
    save_path: Path | None = None,
) -> dict[str, list[str]]:
    """
    Atribui os saldos contados diretamente na planilha original.

    Parâmetros
    ----------
    counted : dict[str, int]
        Dicionário {barcode: quantidade} gerado pelo counter.
    wb : Workbook, opcional
        Workbook já carregado; se None, carrega do caminho padrão.
    save_path : Path, opcional
        Caminho onde salvar; se None, salva no próprio arquivo original.

    Retorna
    -------
    dict com chaves:
        - "matched"     : barcodes encontrados e atualizados
        - "not_found"   : barcodes lidos nos .txt mas ausentes na planilha
    """
    if wb is None:
        wb, original_path = load_workbook()
    else:
        original_path = save_path if save_path is not None else Path(".")

    ws = wb.active
    if ws is None:
        raise ValueError("Workbook não possui uma planilha ativa")

    if save_path is None:
        save_path = original_path

    # Indexa barcode → linha da planilha
    barcode_index = _build_barcode_index(ws, COL_BARCODE, DATA_START_ROW)

    matched: list[str] = []
    not_found: list[str] = []

    for barcode, qty in counted.items():
        row = barcode_index.get(barcode)
        if row is not None:
            ws[f"{COL_QTD_FISICO}{row}"] = qty
            matched.append(barcode)
        else:
            not_found.append(barcode)

    wb.save(save_path)

    # ── Log de resultado ────────────────────────────────────────────────
    print(f"  ✅ Produtos atualizados na planilha: {len(matched)}")

    if not_found:
        print(f"  ⚠️  Barcodes ignorados (não encontrados na planilha): {len(not_found)}")

    print(f"  💾 Planilha atualizada: {save_path}")

    return {"matched": matched, "not_found": not_found}
